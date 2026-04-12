import pandas as pd
from openpyxl import Workbook
import json
import os
import re
import sys


if len(sys.argv) < 2:
    print("Usage: python script.py <excel_file>")
    sys.exit(1)

filename = sys.argv[1]

input_file = f"../Reports/Excel/{filename}"
output_file = f"../Reports/Cleaned/{filename}"

# Bank config file
bank = filename[:3]
config_file = f"Config/{bank}.json"


# LOAD CONFIG
def load_config(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# NORMALIZE TEXT
def normalize_text(text):
    if not isinstance(text, str):
        return ""

    text = text.strip().lower()

    text = re.sub(r"^[^a-zA-Z0-9]+", "", text)
    text = text.replace(".", "")
    text = text.replace("+/-", "")
    text = text.replace("±", "")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


# NUMBER DETECTION
def is_number(s):

    s = str(s).strip()

    test = s.replace(" ", "").replace(",", "")

    if test.startswith("-"):
        test = test[1:]

    if test.startswith("(") and test.endswith(")"):
        test = test[1:-1]

    return test.isdigit()


# FIND INDICATOR (supports multi-line)
def find_indicator(lines, start_index, indicators):

    combined = normalize_text(lines[start_index])

    if combined in indicators:
        return combined, start_index

    j = start_index + 1

    while j < len(lines) and not is_number(lines[j]):

        combined += " " + normalize_text(lines[j])

        if combined in indicators:
            return combined, j

        j += 1

    return None, start_index


# PROCESS SHEET
def process_sheet(lines, sheet_config):

    indicators = sheet_config.get("rows to extract", [])
    standardization = sheet_config.get("standardization", {})

    indicators_norm = {normalize_text(x): x for x in indicators}

    result = []
    i = 0

    if len(lines)>=3:
        result.extend([x for x in lines[:3]])
        i=3

    while i < len(lines):

        indicator_found, last_row = find_indicator(lines, i, indicators_norm)

        if indicator_found:

            original_indicator = indicators_norm[indicator_found]

            # Standardized name
            standardized = standardization.get(original_indicator, original_indicator)

            result.append(standardized)

            numbers = []

            j = last_row + 1

            while j < len(lines) and len(numbers) < 2:

                if is_number(lines[j]):
                    numbers.append(lines[j])

                j += 1

            while len(numbers) < 2:
                numbers.append("0")

            result.extend(numbers)

            i = j

        else:
            i += 1

    return result


# MAIN
config = load_config(config_file)

if filename not in config:
    print(f"No config found for file {filename}")
    sys.exit(1)

file_config = config[filename]


excel = pd.ExcelFile(input_file)
sheet_names = excel.sheet_names

wb = Workbook()
wb.remove(wb.active)

for sheet in sheet_names:

    print(f"Processing sheet: {sheet}")

    sheet_config = file_config.get(sheet)

    if not sheet_config or "rows to extract" not in sheet_config:
        print("  No extraction config → skipping")
        continue

    df = pd.read_excel(input_file, sheet_name=sheet, header=None)

    lines = (
        df.astype(str)
        .stack()
        .dropna()
        .astype(str)
        .str.strip()
        .tolist()
    )

    result = process_sheet(lines, sheet_config)

    ws = wb.create_sheet(title=sheet)

    
    ws.column_dimensions["A"].width = 80

    for idx, val in enumerate(result, start=1):
        ws[f"A{idx}"] = val

    ws["A1"] = "Results"
    print(f"  Extracted {len(result)//3} indicators")


wb.save(output_file)
